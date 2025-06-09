'''
Created on 08-Jun-2025

@author: User1
'''
import openpyxl
from openpyxl import Workbook 

filepath = "D:\\YASHU\\work books\\Book1.xlsx"
my_workbook = openpyxl.load_workbook(filepath)
# my_active_sheet = my_workbook.active # Get the currently active sheet
my_active_sheet = my_workbook["Sheet1"] # Get the sheet by name
total_rows = my_active_sheet.max_row
print(total_rows)

# total_columns = my_active_sheet.max_column
# print(total_columns)
#
# username = my_active_sheet.cell(2, 1).value
# password = my_active_sheet.cell(2, 2).value
# print(username, password)
#
# username = my_active_sheet.cell(3, 1).value
# password = my_active_sheet.cell(3, 2).value
# print(username, password)
#
# username = my_active_sheet.cell(4, 1).value
# password = my_active_sheet.cell(4, 2).value
# print(username, password)
#
# username = my_active_sheet.cell(5, 1).value
# password = my_active_sheet.cell(5, 2).value
# print(username, password)




total_rows = my_active_sheet.max_row
print("Total rows:", total_rows)

total_columns = my_active_sheet.max_column
print("Total columns:", total_columns)

# Loop through rows 2 to total_rows (assuming row 1 has headers)
for i in range(2, total_rows + 1):
    username = my_active_sheet.cell(i, 1).value
    password = my_active_sheet.cell(i, 2).value
    print(username, password)
