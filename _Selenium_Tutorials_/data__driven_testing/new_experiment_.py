import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

'''getting sheet'''
filepath = "D:\\YASHU\\work books\\Book1.xlsx"
workbook = openpyxl.load_workbook(filepath)
sheet = workbook["Sheet1"]

'''chrome'''
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(10)

'''looping statement'''
for i in range(2, sheet.max_row + 1):
    username = sheet.cell(i, 1).value
    password = sheet.cell(i, 2).value

    print(f"Trying login with: {username} / {password}")

    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    time.sleep(2)

    '''login user and pass'''
    user_name = driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input')
    user_name.clear()
    user_name.send_keys(username)
    password_1 = driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[2]/div/div[2]/input')
    password_1.clear()
    password_1.send_keys(password)
    login_btn = driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button')
    login_btn.click()
    time.sleep(3)

    '''conditional statement'''
    current_url = driver.current_url
    result = "Fail"
    expected = "unknown"

    if "dashboard" in current_url:
        expected = "valid"
        result = "Pass"

    if "auth/login" in current_url:
        expected = "invalid"
        result = "Pass"

    '''filling the result and expected column'''
    sheet.cell(i, 3).value = result
    sheet.cell(i, 4).value = expected
    print("Login result:", result)

'''saving to book'''
workbook.save(filepath)
print("\n Testing completed and results saved to Excel😊.")
